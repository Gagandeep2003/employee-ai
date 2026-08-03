"""Generic tabular export helpers for admin reports: CSV, Excel (.xlsx), and PDF, all from
the same (rows, columns) shape -- one implementation of each format shared by every report
endpoint, instead of each reimplementing its own writer.
"""
import csv as csv_module
import io
from typing import Any, List, Tuple

from fastapi.responses import Response

Column = Tuple[str, str]  # (dict key, display label)


def _cell(row: dict, key: str) -> Any:
    v = row.get(key, "")
    return "" if v is None else v


def to_csv_response(rows: List[dict], columns: List[Column], filename: str) -> Response:
    buf = io.StringIO()
    writer = csv_module.writer(buf)
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([_cell(row, key) for key, _ in columns])
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'})


def to_xlsx_response(rows: List[dict], columns: List[Column], filename: str, title: str = "Report") -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]  # Excel sheet names cap at 31 chars
    header_fill = PatternFill(start_color="1E3F33", end_color="1E3F33", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell(row, key))
    for col_idx, (_, label) in enumerate(columns, start=1):
        letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
        ws.column_dimensions[letter].width = max(12, len(label) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def to_pdf_response(rows: List[dict], columns: List[Column], filename: str, title: str = "Report") -> Response:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=15 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 6 * mm)]

    header = [label for _, label in columns]
    data = [header] + [[str(_cell(row, key)) for key, _ in columns] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3F33")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    doc.build(story)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'})


def export_response(fmt: str, rows: List[dict], columns: List[Column], filename: str, title: str = "Report") -> Response:
    fmt = (fmt or "csv").lower()
    if fmt == "csv":
        return to_csv_response(rows, columns, filename)
    if fmt in ("xlsx", "excel"):
        return to_xlsx_response(rows, columns, filename, title)
    if fmt == "pdf":
        return to_pdf_response(rows, columns, filename, title)
    from fastapi import HTTPException
    raise HTTPException(400, f"Unknown export format '{fmt}' -- use csv, xlsx, or pdf")
